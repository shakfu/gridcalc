"""Generate the .xlsx import fixtures in this directory.

The committed ``.xlsx`` files are a persistent corpus for the OpenXLSX-backed
importer (``Grid.xlsxload``). Each file targets a specific category of import
behaviour; ``tests/test_xlsx_fixtures.py`` loads them and asserts on the
result. Loading needs no third-party library, so the importer is testable
even where openpyxl is absent -- only *regenerating* the fixtures needs it.

Run from the repo root to (re)create every fixture::

    uv run python tests/xlsx/generate_fixtures.py

Keep this in sync with the expectations encoded in the test module. The
content here is deliberately deterministic (no dates-of-today, no random
values) so the files round-trip identically on regeneration. No emoji are
used; unicode coverage relies on accents, CJK, Cyrillic, and math symbols.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

HERE = Path(__file__).resolve().parent


def _save(wb: openpyxl.Workbook, name: str) -> None:
    wb.save(str(HERE / name))


def types_xlsx() -> None:
    """Every scalar cell type: int, negative, float, large int, zero, small
    float, string, booleans, and a within-column gap (sparse)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "types"
    ws["A1"] = 42
    ws["A2"] = -7
    ws["A3"] = 3.14159
    ws["A4"] = 1000000
    ws["A5"] = 0
    ws["A6"] = 2.5e-9
    ws["B1"] = "hello"
    ws["B3"] = True  # B2 left blank on purpose (gap)
    ws["B4"] = False
    ws["C1"] = "last"
    _save(wb, "types.xlsx")


def formulas_xlsx() -> None:
    """Formulas: aggregates over a range, cell arithmetic, a text-returning
    IF, a formula over other formulas, and a divide-by-zero error."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "calc"
    for i, v in enumerate([10, 20, 30, 40, 50], start=1):
        ws.cell(row=i, column=1, value=v)  # A1:A5
    ws["B1"] = "=SUM(A1:A5)"
    ws["B2"] = "=AVERAGE(A1:A5)"
    ws["B3"] = "=A1*A2"
    ws["B4"] = '=IF(A1>5, "big", "small")'
    ws["B5"] = "=MAX(A1:A5)"
    ws["B6"] = "=A1/0"
    ws["C1"] = "=B1+B5"  # formula over formulas
    _save(wb, "formulas.xlsx")


def multisheet_xlsx() -> None:
    """Three sheets in workbook order, a sheet name with a space, and a
    cross-sheet formula that reaches both other sheets."""
    wb = openpyxl.Workbook()
    q1 = wb.active
    q1.title = "Q1"
    q1["A1"] = 100
    q1["A2"] = 200
    q2 = wb.create_sheet("Q2")
    q2["A1"] = 150
    q2["A2"] = 250
    rep = wb.create_sheet("Summary Report")
    rep["A1"] = "Total"
    rep["B1"] = "=SUM(Q1!A1:A2)+SUM(Q2!A1:A2)"
    _save(wb, "multisheet.xlsx")


def sparse_xlsx() -> None:
    """Scattered cells with large gaps, plus cells exactly on and one past
    gridcalc's 256-column / 1024-row bounds (the out-of-bounds cells must be
    silently dropped, not crash the import)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sparse"
    ws["A1"] = 1
    ws["E5"] = 5
    ws["J10"] = "deep"
    ws.cell(row=1, column=256, value="col256")  # gridcalc col 255 -- valid
    ws.cell(row=1, column=257, value="dropcol")  # gridcalc col 256 -- dropped
    ws.cell(row=1024, column=1, value="row1024")  # gridcalc row 1023 -- valid
    ws.cell(row=1025, column=1, value="droprow")  # gridcalc row 1024 -- dropped
    _save(wb, "sparse.xlsx")


def text_and_numbers_xlsx() -> None:
    """The text-vs-number interpretation: numeric-looking text becomes a
    number, non-numeric text stays a label, spaces are preserved."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "mix"
    ws["A1"] = "007"  # numeric text -> number 7
    ws["A2"] = "3.5"  # numeric text -> number 3.5
    ws["A3"] = "  padded  "  # spaces preserved -> label
    ws["A4"] = "123abc"  # not fully numeric -> label
    ws["A5"] = "-42"  # signed numeric text -> number -42
    ws["A6"] = 12345678901234  # large integer stays exact
    ws["A7"] = 0.1  # float precision
    _save(wb, "text_and_numbers.xlsx")


def dates_xlsx() -> None:
    """Date / datetime cells import as their Excel serial numbers; a formula
    subtracting two dates yields a day count."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "dates"
    ws["A1"] = dt.date(2024, 1, 1)
    ws["A2"] = dt.datetime(2024, 3, 15, 12, 0, 0)
    ws["A3"] = dt.date(2000, 1, 1)
    ws["B1"] = "=A2-A1"  # 74.5 days
    _save(wb, "dates.xlsx")


def unicode_xlsx() -> None:
    """A unicode sheet name and unicode content (accents, CJK, Cyrillic,
    math symbols); no emoji."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "café"
    ws["A1"] = "café"
    ws["A2"] = "日本語"
    ws["A3"] = "Москва"
    ws["A4"] = "naïve résumé"
    ws["A5"] = "sum: a to b, approx"  # keep ASCII fallback row for a stable anchor
    ws["A6"] = "→ ↔ ≈ ∑"  # arrows + approx + n-ary sum
    _save(wb, "unicode.xlsx")


def empty_xlsx() -> None:
    """A workbook with a single, entirely empty sheet."""
    wb = openpyxl.Workbook()
    wb.active.title = "empty"
    _save(wb, "empty.xlsx")


def named_ranges_xlsx() -> None:
    """Multi-sheet workbook with defined names: a cross-sheet range name, a
    single-cell name, and a same-sheet name, plus a formula-valued name that
    the importer must skip (only simple cell/range refs become named ranges)."""
    wb = openpyxl.Workbook()
    data = wb.active
    data.title = "Data"
    for i, v in enumerate([10, 20, 30, 40], start=1):
        data.cell(row=i, column=1, value=v)  # A1:A4
    data["C1"] = 0.1  # a tax rate, target of a single-cell name
    data["B1"] = "=SUM(Local)"  # same-sheet name -> 30
    report = wb.create_sheet("Report")
    report["A1"] = "=SUM(Nums)"  # cross-sheet range name -> 100
    report["A2"] = "=TaxRate*SUM(Nums)"  # single-cell name -> 10
    report["A3"] = "=SUM(Data!A1:A4)"  # plain-range control -> 100
    wb.defined_names.add(DefinedName("Nums", attr_text="Data!$A$1:$A$4"))
    wb.defined_names.add(DefinedName("TaxRate", attr_text="Data!$C$1"))
    wb.defined_names.add(DefinedName("Local", attr_text="Data!$A$1:$A$2"))
    # A formula-valued name: not a simple reference, so it must NOT import.
    wb.defined_names.add(DefinedName("Total", attr_text="SUM(Data!$A$1:$A$4)"))
    _save(wb, "named_ranges.xlsx")


def table_and_chart_xlsx() -> None:
    """A worksheet carrying an Excel table and an embedded bar chart, plus a
    Report sheet whose formulas use a plain range (works) and a structured
    table reference (unsupported -- documents the gap). The import must not
    choke on the table/chart parts."""
    wb = openpyxl.Workbook()
    sales = wb.active
    sales.title = "Sales"
    sales["A1"] = "Region"
    sales["B1"] = "Amount"
    for i, (region, amt) in enumerate([("North", 100), ("South", 200), ("East", 150)], start=2):
        sales.cell(row=i, column=1, value=region)
        sales.cell(row=i, column=2, value=amt)
    table = Table(displayName="SalesTable", ref="A1:B4")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    sales.add_table(table)
    chart = BarChart()
    chart.add_data(Reference(sales, min_col=2, min_row=1, max_row=4), titles_from_data=True)
    sales.add_chart(chart, "D2")
    report = wb.create_sheet("Report")
    report["A1"] = "=SUM(Sales!B2:B4)"  # plain range -> 450 (control)
    report["A2"] = "=SUM(SalesTable[Amount])"  # structured ref -> unsupported
    _save(wb, "table_and_chart.xlsx")


def main() -> None:
    types_xlsx()
    formulas_xlsx()
    multisheet_xlsx()
    sparse_xlsx()
    text_and_numbers_xlsx()
    dates_xlsx()
    unicode_xlsx()
    empty_xlsx()
    named_ranges_xlsx()
    table_and_chart_xlsx()
    print(f"wrote fixtures to {HERE}")


if __name__ == "__main__":
    main()
